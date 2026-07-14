#!/usr/bin/env python3
"""Import verbatim single-character meanings from Taiwan MOE's open XLSX.

The source is CC BY-ND 3.0 Taiwan. This importer selects matching one-character
entries and normalizes only spreadsheet line-break escapes; it does not
translate, rewrite, or infer meanings from glyph dates.
"""

from __future__ import annotations

import argparse
import collections
import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
METADATA = ROOT / "Resources" / "character_metadata.json"
AUDIT = ROOT / "Audit" / "moe-definition-summary.json"
SOURCE_PAGE = (
    "https://language.moe.gov.tw/001/Upload/Files/site_content/"
    "M0001/respub/dict_reviseddict_download.html"
)


def clean_cell(value: object) -> str | None:
    if not isinstance(value, str) or not value.strip():
        return None
    text = value.replace("_x000D_\r\n", "\n").replace("_x000D_\n", "\n")
    text = text.replace("_x000D_", "\n").replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"\n{3,}", "\n\n", text).strip()
    return text or None


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path)
    parser.add_argument(
        "--wiktionary-cache",
        type=Path,
        help="optional sourced fallback cache created by fetch_wiktionary_definitions.py",
    )
    args = parser.parse_args()

    rows = json.loads(METADATA.read_text(encoding="utf-8"))
    wanted = {row["character"] for row in rows}
    definitions: dict[str, list[tuple[int, str, str]]] = collections.defaultdict(list)

    workbook = load_workbook(args.xlsx, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    header = next(iterator)
    columns = {name: index for index, name in enumerate(header)}
    required = {"字詞名", "字數", "字詞號", "多音排序", "釋義"}
    if not required.issubset(columns):
        raise SystemExit(f"unexpected MOE columns: {header!r}")

    for values in iterator:
        character = values[columns["字詞名"]]
        character_count = values[columns["字數"]]
        if character_count != 1 or character not in wanted:
            continue
        definition = clean_cell(values[columns["釋義"]])
        if not definition:
            continue
        pronunciation_rank = values[columns["多音排序"]]
        entry_number = values[columns["字詞號"]]
        rank = int(pronunciation_rank) if isinstance(pronunciation_rank, int) else 99
        definitions[character].append((rank, str(entry_number or ""), definition))
    workbook.close()

    wiktionary: dict[str, str | None] = {}
    if args.wiktionary_cache and args.wiktionary_cache.is_file():
        wiktionary = json.loads(args.wiktionary_cache.read_text(encoding="utf-8"))

    moe_count = 0
    wiktionary_count = 0
    for row in rows:
        character = row["character"]
        entries = definitions.get(character, [])
        if entries:
            ordered = sorted(entries, key=lambda item: (item[0], item[1]))
            unique: list[str] = []
            for _, _, definition in ordered:
                if definition not in unique:
                    unique.append(definition)
            row["chineseDefinition"] = "\n\n".join(unique)
            row["definitionSource"] = SOURCE_PAGE
            row["definitionAttribution"] = "教育部《重編國語辭典修訂本》"
            moe_count += 1
        elif wiktionary.get(character):
            row["chineseDefinition"] = wiktionary[character]
            row["definitionSource"] = (
                f"https://zh.wiktionary.org/wiki/{character}"
            )
            row["definitionAttribution"] = "中文維基詞典"
            wiktionary_count += 1
        else:
            row["chineseDefinition"] = None
            row["definitionSource"] = None
            row["definitionAttribution"] = None

    METADATA.write_text(
        json.dumps(rows, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    summary = {
        "source": SOURCE_PAGE,
        "source_file": args.xlsx.name,
        "license": "CC BY-ND 3.0 Taiwan",
        "characters": len(rows),
        "with_moe_definition": moe_count,
        "with_wiktionary_fallback": wiktionary_count,
        "with_chinese_definition": moe_count + wiktionary_count,
        "missing_chinese_definition": len(rows) - moe_count - wiktionary_count,
        "transformation": "single-character row selection and line-break normalization only",
    }
    AUDIT.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
